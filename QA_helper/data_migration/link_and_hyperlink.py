import openpyxl
import csv
import sys
import os

def extract_links(xlsx_file, output_file):
    if not os.path.exists(xlsx_file):
        print(f"❌ Error: File '{xlsx_file}' not found.")
        return

    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    sheet = wb["LiveDesign File Import"] if "LiveDesign File Import" in wb.sheetnames else wb.active
    
    automation_data = []
    junk_keywords = ['firefox', 'chrome', 'test case(s)', 'selenium test(s)', 'api test(s)']
    
    for row in sheet.iter_rows(min_row=5):
        test_case_desc = row[2].value
        if not test_case_desc or str(test_case_desc).strip().lower() in junk_keywords: 
            continue
            
        case_clean = str(test_case_desc).strip()

        # Selenium (Col D)
        if row[3].value and str(row[3].value).lower() not in ['n/a', '']:
            link = row[3].hyperlink.target if row[3].hyperlink else ""
            automation_data.append([case_clean, 'Selenium', str(row[3].value).strip(), link])

        # API (Col E)
        if row[4].value and str(row[4].value).lower() not in ['n/a', '']:
            link = row[4].hyperlink.target if row[4].hyperlink else ""
            automation_data.append([case_clean, 'API', str(row[4].value).strip(), link])

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['TestCase', 'Type', 'TestName', 'Link'])
        writer.writerows(automation_data)
        
    print(f"✅ Step 2 Complete: Created {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python 2_extract_automation.py <input_xlsx> <output_file.csv>")
    else:
        extract_links(sys.argv[1], sys.argv[2])